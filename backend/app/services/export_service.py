# Copyright (c) 2025 Savvina AI Ltd
# Licensed under the Business Source License 1.1 — see LICENSE for details.

"""CSV and XLSX generation from stored query results."""

from __future__ import annotations

import contextlib
import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)

_MAX_COL_WIDTH = 50
_MIN_COL_WIDTH = 10
# Cap the number of rows scanned when computing XLSX column auto-width; avoids O(n) cost
# on large result sets where the first 200 rows are representative enough.
_XLSX_WIDTH_SAMPLE_ROWS = 200


def generate_csv(results_json: dict[str, Any]) -> str:
    """Return a CSV string from a QueryResultsResponse-shaped dict."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(results_json.get("columns", []))
    for row in results_json.get("rows", []):
        writer.writerow(row)
    return buf.getvalue()


def generate_xlsx(results_json: dict[str, Any], title: str = "Query Results") -> bytes:
    """Return XLSX bytes with a formatted worksheet.

    Features: bold header, frozen top row, auto-column-width, and numeric
    typing where the column_types hint allows it.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    columns: list[str] = results_json.get("columns", [])
    column_types: list[str] = results_json.get("column_types", [])
    rows: list[list[Any]] = results_json.get("rows", [])

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    wrap_alignment = Alignment(vertical="top", wrap_text=False)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment

    numeric_types = {
        "integer",
        "int",
        "bigint",
        "smallint",
        "float",
        "double",
        "decimal",
        "numeric",
        "real",
        "number",
        "money",
    }

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            if value is None:
                ws.cell(row=row_idx, column=col_idx, value="")
                continue
            ct = column_types[col_idx - 1].lower() if col_idx - 1 < len(column_types) else ""
            if ct in numeric_types:
                with contextlib.suppress(ValueError, TypeError):
                    value = float(value) if "." in str(value) else int(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row in rows[:_XLSX_WIDTH_SAMPLE_ROWS]:
            cell_val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(cell_val or "")))
        width = min(max(max_len + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
