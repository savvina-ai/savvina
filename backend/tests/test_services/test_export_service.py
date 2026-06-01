# Copyright (c) 2025 Savvina AI Ltd
# Licensed under the Business Source License 1.1 — see LICENSE for details.

"""Tests for export_service — CSV and XLSX generation from query results."""

from __future__ import annotations

import csv
import io

from app.services.export_service import _MAX_COL_WIDTH, _MIN_COL_WIDTH, generate_csv, generate_xlsx


def _results(columns, column_types, rows):
    return {"columns": columns, "column_types": column_types, "rows": rows}


# ── CSV tests ─────────────────────────────────────────────────────────────────


def test_csv_basic():
    data = _results(["id", "name"], ["integer", "text"], [[1, "Alice"], [2, "Bob"]])
    out = generate_csv(data)
    reader = list(csv.reader(io.StringIO(out)))
    assert reader[0] == ["id", "name"]
    assert reader[1] == ["1", "Alice"]
    assert reader[2] == ["2", "Bob"]


def test_csv_null_values():
    data = _results(["a", "b"], ["text", "text"], [[None, "x"], ["y", None]])
    out = generate_csv(data)
    reader = list(csv.reader(io.StringIO(out)))
    assert reader[1] == ["", "x"]
    assert reader[2] == ["y", ""]


def test_csv_delimiter_in_value():
    data = _results(["text"], ["text"], [['say "hello", world']])
    out = generate_csv(data)
    reader = list(csv.reader(io.StringIO(out)))
    assert reader[1] == ['say "hello", world']


def test_csv_unicode():
    data = _results(["city"], ["text"], [["Zürich"], ["東京"], ["São Paulo"]])
    out = generate_csv(data)
    assert "Zürich" in out
    assert "東京" in out
    assert "São Paulo" in out


def test_csv_empty_results():
    data = _results(["col1", "col2"], ["text", "text"], [])
    out = generate_csv(data)
    reader = list(csv.reader(io.StringIO(out)))
    assert reader[0] == ["col1", "col2"]
    assert len(reader) == 1


# ── XLSX tests ────────────────────────────────────────────────────────────────


def _read_xlsx(raw: bytes):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    return ws


def test_xlsx_basic():
    data = _results(["id", "name"], ["integer", "text"], [[1, "Alice"], [2, "Bob"]])
    raw = generate_xlsx(data)
    ws = _read_xlsx(raw)
    assert ws.cell(1, 1).value == "id"
    assert ws.cell(1, 2).value == "name"
    assert ws.cell(2, 1).value == 1
    assert ws.cell(2, 2).value == "Alice"


def test_xlsx_empty_results():
    data = _results(["col1", "col2"], [], [])
    raw = generate_xlsx(data)
    ws = _read_xlsx(raw)
    assert ws.cell(1, 1).value == "col1"
    assert ws.cell(1, 2).value == "col2"
    assert ws.max_row == 1


def test_xlsx_null_values_produce_empty_or_none_cell():
    # openpyxl reads back "" as None; either way the cell must be falsy
    data = _results(["a"], ["text"], [[None]])
    ws = _read_xlsx(generate_xlsx(data))
    assert not ws.cell(2, 1).value


def test_xlsx_numeric_typing():
    data = _results(["price", "qty"], ["decimal", "integer"], [["9.99", "3"]])
    ws = _read_xlsx(generate_xlsx(data))
    assert ws.cell(2, 1).value == 9.99
    assert ws.cell(2, 2).value == 3


def test_xlsx_column_width_capped():
    long_value = "x" * 200
    data = _results(["col"], ["text"], [[long_value]])
    raw = generate_xlsx(data)
    ws = _read_xlsx(raw)
    from openpyxl.utils import get_column_letter

    width = ws.column_dimensions[get_column_letter(1)].width
    assert width <= _MAX_COL_WIDTH


def test_xlsx_column_width_floor():
    data = _results(["c"], ["text"], [["a"]])
    raw = generate_xlsx(data)
    ws = _read_xlsx(raw)
    from openpyxl.utils import get_column_letter

    width = ws.column_dimensions[get_column_letter(1)].width
    assert width >= _MIN_COL_WIDTH


def test_xlsx_samples_only_first_200_rows():
    """Width computation samples at most 200 rows; 201st row should not inflate width."""
    short_val = "x"
    long_val = "y" * 100
    rows = [[short_val]] * 200 + [[long_val]]
    data = _results(["col"], ["text"], rows)
    raw = generate_xlsx(data)
    ws = _read_xlsx(raw)
    from openpyxl.utils import get_column_letter

    # With 200-row cap, long_val at row 201 should not affect column width
    width = ws.column_dimensions[get_column_letter(1)].width
    expected_max = min(max(len(short_val) + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
    assert width <= expected_max
